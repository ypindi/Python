import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    # wb = xl.load_workbook('transactions.xlsx')
    # returns a workbook object, which we are storing in wb
    sheet = wb['Sheet1']

    # cell = sheet['a1']
    # cell = sheet(1,1)
    # same thing
    # print(cell.value)

    # print(sheet.max_row)
    # total rows

    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        # print(cell)
        print(cell.value)
        correct_price = cell.value*0.9
        correct_price_cell = sheet.cell(row, 4)
        correct_price_cell.value = correct_price

    # values object will have D2-D4 
    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # wb.save('transactions2.xlsx')
    wb.save(filename)